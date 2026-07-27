#!/usr/bin/env python3
"""Assemble NNDR ratepayer presence evidence from the per-council files in
~/observatory-data/raw/nndr/. Schemas vary; columns are located by header
heuristics. Output is an EVIDENCE map (normalised ratepayer name -> councils,
postcodes, company numbers where published), used to upgrade nonLocal
suppliers to tradingExternal with "business-rates ratepayer" evidence.
The raw data is NOT republished (several councils state no licence; used as
evidence only, disclosed on the method page).
"""
import csv, json, sys
from collections import defaultdict
from pathlib import Path
import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from resolve_suppliers import normalise

RAW = Path.home() / "observatory-data/raw/nndr"
OUT = Path.home() / "observatory-data/processed/nndr_presence.json"
COUNCILS = {"blackburn": "Blackburn with Darwen", "blackpool": "Blackpool",
            "burnley": "Burnley", "chorley": "Chorley", "fylde": "Fylde",
            "lancaster": "Lancaster", "pendle": "Pendle", "preston": "Preston",
            "ribble_valley": "Ribble Valley", "rossendale": "Rossendale",
            "south_ribble": "South Ribble", "west_lancashire": "West Lancashire"}

NAME_HDRS = ("account holder", "account holder1", "lead_liable", "liable party",
             "primary liable party name", "company name", "account name",
             "ratepayer", "primary liable party")
CRN_HDRS = ("company_number", "company number")
PC_HDRS = ("postcode", "post code")

def header_index(headers):
    hl = [str(h or "").strip().lower() for h in headers]
    def find(cands):
        for c in cands:
            for i, h in enumerate(hl):
                if h == c or h.startswith(c):
                    return i
        return None
    return find(NAME_HDRS), find(CRN_HDRS), find(PC_HDRS), hl

def rows_from(path):
    if path.suffix == ".csv":
        with open(path, newline="", encoding="utf-8", errors="replace") as f:
            yield from csv.reader(f)
    else:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(values_only=True):
            yield list(row)

presence = defaultdict(lambda: {"councils": set(), "postcodes": set(),
                                "crns": set(), "records": 0})
report = {}
for slug, lad in COUNCILS.items():
    matches = sorted(RAW.glob(f"{slug}.*"))
    if not matches:
        report[slug] = "no file"
        continue
    path = matches[0]
    ni = ci = pi = None
    n = 0
    try:
        for row in rows_from(path):
            if ni is None:
                f_ni, f_ci, f_pi, hl = header_index(row)
                if f_ni is not None:
                    ni, ci, pi = f_ni, f_ci, f_pi
                continue
            if ni >= len(row):
                continue
            name = str(row[ni] or "").strip()
            if not name or len(name) < 3:
                continue
            key = normalise(name)
            if not key:
                continue
            p = presence[key]
            p["councils"].add(lad)
            p["records"] += 1
            if pi is not None and pi < len(row) and row[pi]:
                p["postcodes"].add(str(row[pi]).strip().upper())
            if ci is not None and ci < len(row) and row[ci]:
                crn = str(row[ci]).strip().upper().replace(" ", "")
                if crn.isdigit():
                    crn = crn.zfill(8)
                if 6 <= len(crn) <= 8:
                    p["crns"].add(crn)
            n += 1
    except Exception as e:
        report[slug] = f"parse error: {e}"
        continue
    report[slug] = f"{n} ratepayer rows"

out = {"$meta": {"retrieved": "2026-07-27",
                 "note": "Evidence map only; raw NNDR data not republished. "
                         "12 of 14 billing authorities publish ratepayer-level "
                         "data (Hyndburn withholds; Wyre publishes without "
                         "names).",
                 "perCouncil": report},
       "byName": {k: {"councils": sorted(v["councils"]),
                      "postcodes": sorted(v["postcodes"])[:20],
                      "crns": sorted(v["crns"]), "records": v["records"]}
                  for k, v in presence.items()}}
OUT.write_text(json.dumps(out))
print(json.dumps(report, indent=1))
print(f"presence map: {len(presence)} distinct ratepayer names")
crn_count = sum(1 for v in presence.values() if v["crns"])
print(f"with company numbers: {crn_count}")
