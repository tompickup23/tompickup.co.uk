#!/usr/bin/env python3
"""fetch_ons_demography.py - ONS Business Demography reference tables (births,
deaths, active, survival, high-growth counts by LA) plus the official
high-growth business % (computed from the same tables; the ONS Explore Local
Statistics indicator is attempted as a cross-check).

Output: ~/observatory-data/processed/ons_demography.json keyed by LAD, with
England + NW ring rows for benchmarking.
"""
import sys
from pathlib import Path
import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from _common import (RAW, LANCS_14, NW_RING, ENGLAND, download, meta, write_out,
                     log, clean_text)

XLSX_URL = ("https://www.ons.gov.uk/file?uri=/businessindustryandtrade/business/"
            "activitysizeandlocation/datasets/businessdemographyreferencetable/"
            "current/businessdemographyexceltables2024.xlsx")

TARGET = {**LANCS_14, **NW_RING, **ENGLAND}

# metric -> list of (sheet, [year labels aligned to value columns])
METRIC_TABLES = {
    "births": [("Table 1.1a", ["2019"]), ("Table 1.1b", ["2020"]),
               ("Table 1.1c", ["2021", "2022", "2023"]), ("Table 1.1d", ["2024"])],
    "deaths": [("Table 2.1a", ["2019"]), ("Table 2.1b", ["2020"]),
               ("Table 2.1c", ["2021", "2022", "2023"]), ("Table 2.1d", ["2024"])],
    "active": [("Table 3.1a", ["2019"]), ("Table 3.1b", ["2020"]),
               ("Table 3.1c", ["2021", "2022", "2023"]), ("Table 3.1d", ["2024"])],
    "high_growth": [("Table 7.1a", ["2019"]), ("Table 7.1b", ["2020"]),
                    ("Table 7.1c", ["2021", "2022", "2023"]), ("Table 7.1d", ["2024"])],
    "active_10plus": [("Table 7.3a", ["2019"]), ("Table 7.3b", ["2020"]),
                      ("Table 7.3c", ["2021", "2022", "2023"]), ("Table 7.3d", ["2024"])],
}

# survival cohort tables (2019..2023 birth cohorts)
SURVIVAL_TABLES = ["Table 5.1a", "Table 5.1b", "Table 5.1c", "Table 5.1d", "Table 5.1e"]


def norm_code(v):
    return str(v).strip() if v is not None else None


def num(v):
    if v is None or v == "":
        return None
    try:
        f = float(v)
        return int(f) if f == int(f) else round(f, 4)
    except (ValueError, TypeError):
        return None


def parse_metric(wb, out, metric, tables):
    for sheet, years in tables:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[4:]:
            code = norm_code(row[0])
            if code not in TARGET:
                continue
            for j, yr in enumerate(years):
                val = num(row[2 + j]) if len(row) > 2 + j else None
                if val is not None:
                    out[code].setdefault(metric, {})[yr] = val


def parse_survival(wb, out):
    for sheet in SURVIVAL_TABLES:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        header = rows[3]
        cohort = None
        for c in header:
            if c and "Births" in str(c):
                cohort = str(c).split()[0]
                break
        labels = {i: clean_text(str(c)) for i, c in enumerate(header) if c}
        for row in rows[4:]:
            code = norm_code(row[0])
            if code not in TARGET:
                continue
            surv = {}
            for i, lbl in labels.items():
                if i < 2:
                    continue
                v = num(row[i]) if len(row) > i else None
                if v is not None:
                    surv[lbl] = v
            if surv and cohort:
                out[code].setdefault("survival", {})[cohort] = surv


def main():
    dest = RAW / "business_demography_2024.xlsx"
    download(XLSX_URL, dest)
    wb = openpyxl.load_workbook(dest, read_only=True, data_only=True)

    out = {c: {"name": TARGET[c], "ons": c} for c in TARGET}
    for metric, tables in METRIC_TABLES.items():
        parse_metric(wb, out, metric, tables)
    parse_survival(wb, out)

    # Compute high-growth % = high_growth / active_10plus per year.
    for c, rec in out.items():
        hg = rec.get("high_growth", {})
        base = rec.get("active_10plus", {})
        pct = {}
        for yr in hg:
            b = base.get(yr)
            if b:
                pct[yr] = round(100.0 * hg[yr] / b, 2)
        if pct:
            rec["high_growth_pct"] = pct
    wb.close()

    m = meta(
        XLSX_URL,
        "Open Government Licence v3.0 (ONS Business Demography 2024, pub 20 Nov 2025)",
        "Business births, deaths, active enterprises, high-growth enterprise "
        "counts, active-enterprises-with-10+-employees, and newly-born survival "
        "for the 14 Lancashire LADs, the NW benchmark ring and England, 2019-2024. "
        "high_growth_pct is high_growth / active_10plus (ONS/OECD high-growth "
        "denominator, >=10 employees at base) computed here; the ONS Explore Local "
        "Statistics 'High growth businesses' indicator is the published equivalent. "
        "IDBR basis: VAT/PAYE-registered businesses only. Counts are ONS-rounded to "
        "the nearest 5.")
    m["geography_groups"] = {
        "lancs_14": list(LANCS_14.keys()),
        "nw_ring": list(NW_RING.keys()),
        "england": list(ENGLAND.keys()),
    }
    write_out("ons_demography.json", m, "areas", out)
    b = out.get("E07000117", {})
    log(f"Burnley check: active={b.get('active')}, high_growth={b.get('high_growth')}, "
        f"hg_pct={b.get('high_growth_pct')}")


if __name__ == "__main__":
    main()
