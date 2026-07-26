#!/usr/bin/env python3
"""fetch_innovate.py - Innovate UK funded projects since 2004.

Scrapes the UKRI publication page for the two current xlsx files, downloads
both, filters rows where Address Local Authority is one of the 14 Lancs LADs.
Keeps ALL Lancashire rows including KTPs. Also records a NW-region aggregate
count. Output: ~/observatory-data/processed/innovate_lancs.json
"""
import sys
import re
from pathlib import Path
import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from _common import (RAW, LANCS_14, download, get_json, meta, write_out,
                     clean_text, log, UA, fresh)

PAGE = "https://www.ukri.org/publications/innovate-uk-funded-projects-since-2004/"


def norm(s):
    return re.sub(r"[^a-z]", "", (s or "").lower())


LA_TO_LAD = {norm(v): k for k, v in LANCS_14.items()}
# Innovate uses "Blackburn with Darwen" and standard LAD names.


def discover_links():
    import requests
    r = requests.get(PAGE, headers={"User-Agent": UA}, timeout=60)
    r.raise_for_status()
    links = re.findall(r'href="([^"]+\.xlsx[^"]*)"', r.text, re.I)
    links = sorted(set(links))
    log(f"found {len(links)} xlsx links")
    return links


def parse_file(path, rows, nw_counter):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    hdr = list(next(it))
    idx = {h: i for i, h in enumerate(hdr)}

    def col(row, name):
        i = idx.get(name)
        return row[i] if i is not None and i < len(row) else None

    n_total = 0
    for row in it:
        n_total += 1
        la = col(row, "Address Local Authority")
        region = col(row, "Address Region")
        if region and "north west" in str(region).lower():
            nw_counter[0] += 1
        lad = LA_TO_LAD.get(norm(la))
        if not lad:
            continue
        rows.append({
            "participant": clean_text(col(row, "Participant Name")),
            "crn": (str(col(row, "CRN")).strip() if col(row, "CRN") else None),
            "project_title": clean_text(col(row, "Project Title")),
            "competition_year": clean_text(col(row, "Competition Year")),
            "product_type": clean_text(col(row, "Innovate UK Product Type")),
            "award_offered": col(row, "Award Offered (£)"),
            "total_costs": col(row, "Total Costs (£)"),
            "status": clean_text(col(row, "Project Status")),
            "lad": lad,
            "postcode": (str(col(row, "Postcode")).strip() if col(row, "Postcode") else None),
            "is_lead": (str(col(row, "Is Lead Participant")).strip().lower() == "yes"),
        })
    wb.close()
    log(f"  {path.name}: scanned {n_total} rows")


def main():
    links = discover_links()
    rows = []
    nw_counter = [0]
    for url in links:
        fname = url.split("/")[-1].split("?")[0]
        dest = RAW / fname
        download(url, dest)
        parse_file(dest, rows, nw_counter)

    per_lad = {}
    ktp = 0
    for r in rows:
        per_lad[r["lad"]] = per_lad.get(r["lad"], 0) + 1
        if r["product_type"] and "knowledge transfer" in r["product_type"].lower():
            ktp += 1

    m = meta(
        PAGE,
        "Open Government Licence v3.0 (UKRI / Innovate UK)",
        "All participation rows where Address Local Authority is one of the 14 "
        "Lancashire LADs, from both bulk xlsx files (2004-2015/16 and 2016/17 to "
        f"present). Includes KTPs ({ktp} rows). CRN present on most rows. "
        f"NW-region aggregate participation count: {nw_counter[0]}.")
    m["per_lad_counts"] = {LANCS_14[k]: v for k, v in per_lad.items()}
    m["nw_region_row_count"] = nw_counter[0]
    m["ktp_rows"] = ktp
    write_out("innovate_lancs.json", m, "projects", rows)


if __name__ == "__main__":
    main()
